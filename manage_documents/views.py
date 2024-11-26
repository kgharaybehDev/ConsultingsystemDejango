import datetime
import os
import logging
from html.entities import html5

from django.contrib.staticfiles import finders
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from django_countries.fields import Country
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image, ListFlowable, ListItem,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from candidates.models import Candidate, Experience
from django.templatetags.static import static
from django.core.files.storage import default_storage
logger = logging.getLogger(__name__)


from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Image
from django.contrib.staticfiles import finders
import os
import logging

logger = logging.getLogger(__name__)

def add_page_frame(canvas, doc):
    canvas.saveState()

    # Set RGB color for rectangles
    red, green, blue = 139 / 255, 13 / 255, 53 / 255

    # Outer rectangle (line width reduced by 50%)
    canvas.setStrokeColorRGB(red, green, blue)
    canvas.setLineWidth(1)
    x0, y0 = 0.5 * inch, 0.5 * inch
    width0, height0 = A4[0] - 1 * inch, A4[1] - 1 * inch
    canvas.rect(x0, y0, width0, height0, stroke=1, fill=0)

    # Middle rectangle with reduced spacing
    canvas.setLineWidth(2)
    x1, y1 = x0 + 0.05 * inch, y0 + 0.05 * inch  # Reduced from 0.1 * inch to 0.05 * inch
    width1, height1 = width0 - 0.1 * inch, height0 - 0.1 * inch  # Reduced from 0.2 * inch to 0.1 * inch
    canvas.rect(x1, y1, width1, height1, stroke=1, fill=0)

    # Inner rectangle with reduced spacing
    canvas.setLineWidth(1)
    x2, y2 = x1 + 0.05 * inch, y1 + 0.05 * inch  # Reduced from 0.1 * inch to 0.05 * inch
    width2, height2 = width1 - 0.1 * inch, height1 - 0.1 * inch  # Reduced from 0.2 * inch to 0.1 * inch
    canvas.rect(x2, y2, width2, height2, stroke=1, fill=0)

    # Draw logo if available
    logo_path = finders.find('images/cv_log.png')
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path, width=1 * inch, height=1 * inch)
        img.drawOn(canvas, A4[0] / 2 - (0.4 * inch), A4[1] - 2 * inch)
    else:
        logger.error("Logo not found at the specified path.")

    canvas.restoreState()


def format_date(date, default='N/A'):
    return date.strftime('%d/%b/%Y') if date else default


def get_photo_url(candidate):
    if candidate.personal_image and hasattr(candidate.personal_image, 'url'):
        return candidate.personal_image.url
    else:
        photo_url = static('images/avatar.png')
        if default_storage.exists(photo_url):
            return photo_url
        else:
            logger.error("Default avatar not found. Using placeholder text.")
            return None


def get_related_values(obj, attr_name):
    try:
        return list(getattr(obj, attr_name).all())
    except AttributeError:
        return []


def create_info_table(title, entries, styles_dict, colWidths):
    bold_blue_style = styles_dict['bold_blue_style']
    regular_style = styles_dict['regular_style']

    data = [
        [
            Paragraph(f"<b>{title}</b>", bold_blue_style),
            Paragraph(entries, regular_style)
        ]
    ]

    table = Table(data, colWidths=colWidths, splitInRow=1)

    table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.blue),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    return table


def get_candidate_info_table(candidate_data, styles_dict):
    bold_blue_style = styles_dict['bold_blue_style']
    regular_style = styles_dict['regular_style']

    # Content for the first column (Candidate Information)
    col0_content = [
        Paragraph(candidate_data['name'].upper(), bold_blue_style),
        Spacer(1, 12),
        Paragraph(f"Email: {candidate_data['email']}", regular_style),
        Paragraph(f"Phone: {candidate_data['mobile']}", regular_style),
    ]

    # Content for the second column (Candidate Photo)
    if candidate_data['photo_path']:
        photo = Image(candidate_data['photo_path'], width=1 * inch, height=1 * inch)
    else:
        photo = Paragraph("No Image Available", regular_style)

    # Define table data with two columns
    data = [
        [col0_content, photo]
    ]

    # Define column widths
    table = Table(data, colWidths=[(0.7 * A4[0] - (1 * inch)), (0.3 * A4[0] - (0.8 * inch))])

    # Apply table styles
    table.setStyle(TableStyle([
        # Align all cells vertically to the top
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),

        # Remove default padding
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),

        # Style for the first column (Candidate Information)
        ('TEXTCOLOR', (0, 0), (0, 0), colors.blue),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),

        # Add a border around the image cell (second column)
        ('BOX', (1, 0), (1, 0), 1, colors.black),  # (start_col, start_row), (end_col, end_row), line width, color
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
        # Optionally, add inner grid lines to the image cell if desired
        ('INNERGRID', (1, 0), (1, 0), 0.5, colors.gray),

        # Add padding inside the image cell for better spacing
        ('TOPPADDING', (1, 0), (1, 0), 5),
        ('BOTTOMPADDING', (1, 0), (1, 0), 5),
        ('LEFTPADDING', (1, 0), (1, 0), 5),
        ('RIGHTPADDING', (1, 0), (1, 0), 5),
    ]))

    return table


def get_section_table(section_name, candidate_data, key, styles_dict, colWidths, item_template):
    items = candidate_data.get(key, [])
    if items:
        entries = "".join(item_template(item) for item in items)
    else:
        entries = "N/A"

    return create_info_table(section_name, entries, styles_dict, colWidths)


def get_declaration_table(candidate_data, styles_dict, colWidths):
    bold_blue_style = styles_dict['bold_blue_style']
    regular_style = styles_dict['regular_style']
    date = datetime.date.today().strftime("%d/%m/%Y")
    signature = candidate_data['name']

    declaration_text = (
        "I hereby declare the above mentioned information is true and verifiable to the best of my knowledge and "
        "I bear responsibility for the correctness of the above mentioned particulars.<br/><br/>"
        f"<b>Date</b>: {date}  &nbsp; &nbsp; &nbsp; <b>Signature:</b> {signature}"
    )

    data = [
        [
            Paragraph("<b>Declaration</b>", bold_blue_style),
            Paragraph(declaration_text, regular_style)
        ]
    ]

    table = Table(data, colWidths=colWidths)

    table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.blue),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    return table


def getCountryName(code=None):
    if code:
        return Country(code=code).name
    else:
        return None


def clone_html(html):
    html = str(html)
    html = html.replace("<p>", "-")
    html = html.replace("</p>", "<br/>")
    html = html.replace("<br>", "<br/>")  # This line ensures all <br> tags are self-closing
    html = html.replace("&nbsp;", " ")
    # Do not wrap in <ul> tags here
    return html


def getResponsibilities(experience):
    responsibilities = (getattr(experience, 'job_responsibilities', ''))
    if len(responsibilities) > 15:
        return  f"Responsibilities:<br/>  <ul>"+ "".join(clone_html(getattr(experience, 'job_responsibilities', '')))+ "</ul>"
    else:
        return f" "


def candidate_export_pdf_CV(request, pk):
    # Get the candidate instance or return 404 if not found
    candidate = get_object_or_404(Candidate, pk=pk)

    # Collecting candidate data
    candidate_data = {
        'name': candidate.full_name,
        'mobile': candidate.call_phone_number,
        'email': candidate.email,
        'photo_path': get_photo_url(candidate),
        'education': get_related_values(candidate, 'educations'),
        'clinical_experience': get_related_values(candidate, 'experiences'),
        'license': get_related_values(candidate, 'licenses'),
        'training_courses': get_related_values(candidate, 'training_courses'),
        'internships': get_related_values(candidate, 'internships'),
        'publications': get_related_values(candidate, 'publications'),
        # 'references': get_related_values(candidate, 'references'),  # Removed as references are within experiences
    }

    # Collect references from experiences
    references = []
    for experience in candidate_data['clinical_experience']:
        if experience.reference_name:
            references.append({
                'name': experience.reference_name,
                'position': experience.reference_job_title,
                'contact_info': experience.reference_contact_info,
                'company': experience.company_name,
            })
    candidate_data['references'] = references

    # Create the HTTP response with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{candidate.full_name}_CV.pdf"'

    # Create the PDF object, using the response object as its "file."
    doc = SimpleDocTemplate(response, pagesize=A4)

    # Define styles
    styles = getSampleStyleSheet()
    bold_blue_style = ParagraphStyle(
        name="BoldBlue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        textColor=colors.blue,
        fontSize=12,
    )
    regular_style = styles["BodyText"]

    styles_dict = {
        'bold_blue_style': bold_blue_style,
        'regular_style': regular_style
    }

    colWidths = [(0.3 * A4[0] - (0.8 * inch)), (0.7 * A4[0] - (1 * inch))]

    elements = []

    doc.leftMargin = 1.5 * inch
    doc.rightMargin = 1.5 * inch
    doc.topMargin = 2.5 * inch
    doc.bottomMargin = 2.5 * inch
    doc.title = f"{candidate.full_name}'s CV"

    # Add elements to the document
    elements.append(get_candidate_info_table(candidate_data, styles_dict))
    elements.append(Spacer(1, 12))

    # List of sections to include
    sections = [
        ('Educational Qualifications', 'education', lambda edu: (
            f"{getattr(edu, 'degree', 'N/A') or 'N/A'} - {getattr(edu, 'field_of_study', 'N/A') or 'N/A'}<br/>"
            f"{getattr(edu, 'institution', 'N/A') or 'N/A'}<br/>"
            f"({format_date(getattr(edu, 'start_date', None))} - {format_date(getattr(edu, 'end_date', None), default='Present')})<br/><br/>"
        )),
        ('Internship', 'internships', lambda internship: (
            f"{getattr(internship, 'company_name', 'N/A') or 'N/A'} , {getattr(internship, 'company_location', 'N/A') or 'N/A'}<br/>"
            f"{getattr(internship, 'job_title', 'N/A') or 'N/A'}<br/>"
            f"({format_date(getattr(internship, 'start_date', None))} - {format_date(getattr(internship, 'end_date', None), default='Present')})<br/><br/>"
        )),
        ('Clinical Experience (including training)', 'clinical_experience', lambda experience: (
                f"{getattr(experience, 'company_name', 'N/A') or 'N/A'} , {getCountryName(code=getattr(experience, 'company_location')) or 'N/A'}<br/>"
                f"{getattr(experience, 'job_title', 'N/A') or 'N/A'}<br/>"
                f"({format_date(getattr(experience, 'start_date', None))} - {format_date(getattr(experience, 'end_date', None), default='Present')})<br/>"
                + getResponsibilities(experience)
                + "<br/>"
        )),

        ('Licenses', 'license', lambda lic: (
            f"{getattr(lic, 'license_name', 'N/A') or 'N/A'} from {getattr(lic, 'license_provider', 'N/A') or 'N/A'}<br/>"
            f"Issue Date: {format_date(getattr(lic, 'issued_date', None))}<br/>"
            f"Expiry Date: {format_date(getattr(lic, 'expiry_date', None))}<br/><br/>"
        )),
        ('Training Courses', 'training_courses', lambda course: (
            f"{getattr(course, 'course_name', 'N/A') or 'N/A'}<br/>"
            f" {getCountryName(code=getattr(course, 'location')) or 'N/A'}<br/>"
            f"{format_date(getattr(course, 'end_date', None))}<br/><br/>"
        )),
        ('Publications', 'publications', lambda pub: (
            f"{getattr(pub, 'title', 'N/A') or 'N/A'}<br/>"
            f"Published in: {getattr(pub, 'journal', 'N/A') or 'N/A'}<br/>"
            f"Date: {format_date(getattr(pub, 'date', None))}<br/><br/>"
        )),
        ('References', 'references', lambda ref: (
            # f"{ref.get('name', 'N/A') or 'N/A'}<br/>"
            f"{ref.get('company', 'N/A') or 'N/A'}, {ref.get('position', 'N/A') or 'N/A'}<br/>"
            f"Contact: {ref.get('contact_info', 'N/A') or 'N/A'}<br/><br/>"
        )),
    ]

    for title, key, item_template in sections:
        elements.append(get_section_table(title, candidate_data, key, styles_dict, colWidths, item_template))
        elements.append(Spacer(1, 12))

    elements.append(get_declaration_table(candidate_data, styles_dict, colWidths))

    # Build the PDF
    doc.build(elements, onFirstPage=add_page_frame, onLaterPages=add_page_frame)

    return response
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import tempfile

def HMC_Sheet(request, pk):
    candidate = get_object_or_404(Candidate, pk=pk)

    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Candidate Data"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    value_font = Font(size=11)
    alignment = Alignment(horizontal="left", vertical="center")

    def append_with_style(sheet, row_data, is_header=False):
        row = sheet.max_row + 1
        for col, value in enumerate(row_data, start=1):
            if isinstance(value, dict) and value.get("link"):  # Check if the value is a link dictionary
                cell = sheet.cell(row=row, column=col, value=value.get("text"))
                sheet.cell(row=row, column=col).hyperlink = value.get("link")
                sheet.cell(row=row, column=col).font = Font(color="0000FF", underline="single")
            else:
                cell = sheet.cell(row=row, column=col, value=value)
            if is_header:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = value_font
            cell.alignment = alignment

    # Table 1: Candidate Basic Information
    append_with_style(sheet, ["Subject", "Value"], is_header=True)
    append_with_style(sheet, ["Email", candidate.email])
    append_with_style(sheet, ["Name", candidate.full_name])
    append_with_style(sheet, ["National ID Number", candidate.national_id_number])
    append_with_style(sheet, ["Phone Number", candidate.call_phone_number or candidate.whatsapp_phone_number])

    # Add a separator row
    sheet.append([])

    # Table 2: Candidate Personal Information
    append_with_style(sheet, ["Subject", "Value"], is_header=True)
    append_with_style(sheet, ["Name", candidate.full_name])
    append_with_style(sheet, ["Birth Date", candidate.birthday])
    append_with_style(sheet, ["Gender", candidate.gender])
    append_with_style(sheet, ["Passport Copy", {"text": "Download", "link": candidate.passport_copy.url} if candidate.passport_copy else "Document does not exist"])
    append_with_style(sheet, ["Passport Expiration Date", candidate.passport_expiration_date])
    append_with_style(sheet, ["Passport Number", candidate.passport_id])
    append_with_style(sheet, ["Personal Image", {"text": "Download", "link": candidate.personal_image.url} if candidate.personal_image else "Document does not exist"])
    append_with_style(sheet, ["National ID Number", candidate.national_id_number])
    append_with_style(sheet, ["Phone Number", candidate.call_phone_number or candidate.whatsapp_phone_number])
    append_with_style(sheet, ["Address", candidate.address])
    append_with_style(sheet, ["ID Copy", {"text": "Download", "link": candidate.national_id_copy.url} if candidate.national_id_copy else "Document does not exist"])

    # Add a separator row
    sheet.append([])

    # Table 3: Candidate Education Information
    for education in candidate.educations.all():
        append_with_style(sheet, ["Subject", "Value"], is_header=True)
        append_with_style(sheet, ["Degree", education.degree.degree])
        append_with_style(sheet, ["Institution Country", education.institution.country.name if education.institution.country else None])
        append_with_style(sheet, ["Institution Name", education.institution.institution])
        append_with_style(sheet, ["Field of Study", education.field_of_study.field_of_study])
        append_with_style(sheet, ["Start Date", education.start_date])
        append_with_style(sheet, ["End Date", education.end_date])
        append_with_style(sheet, ["Certification Copy", {"text": "Download", "link": education.certification_copy.url} if education.certification_copy else "Document does not exist"])
        append_with_style(sheet, ["Transcript Copy", {"text": "Download", "link": education.transcript_copy.url} if education.transcript_copy else "Document does not exist"])
        sheet.append([])  # Add a separator row between educations

    # Table 4: Candidate Experience Information
    for experience in candidate.experiences.all():
        append_with_style(sheet, ["Subject", "Value"], is_header=True)
        append_with_style(sheet, ["Start Date", experience.start_date])
        append_with_style(sheet, ["End Date", experience.end_date])
        append_with_style(sheet, ["Job Title", experience.job_title])
        append_with_style(sheet, ["Company Name", experience.company_name])
        append_with_style(sheet, ["Company Location", experience.get_company_location()])
        sheet.append([])  # Add a separator row between experiences

    # Table 5: Candidate Licenses Information
    for license in candidate.licenses.all():
        append_with_style(sheet, ["Subject", "Value"], is_header=True)
        append_with_style(sheet, ["License Name", license.license_name])
        append_with_style(sheet, ["License Provider Country", license.license_provider.country.name if license.license_provider.country else None])
        append_with_style(sheet, ["License Provider Name", license.license_provider.name])
        append_with_style(sheet, ["License Number", license.license_number])
        append_with_style(sheet, ["License Status", "Expired" if license.is_expired() else "Valid"])
        append_with_style(sheet, ["Start Date", license.issued_date])
        append_with_style(sheet, ["End Date", license.expiry_date])
        sheet.append([])  # Add a separator row between licenses

    # Table 6: Candidate Training Courses Information
    for course in candidate.training_courses.all():
        append_with_style(sheet, ["Subject", "Value"], is_header=True)
        append_with_style(sheet, ["Course Name", course.course_name])
        append_with_style(sheet, ["Location", course.location.name if course.location else None])
        append_with_style(sheet, ["End Date", course.end_date])
        sheet.append([])  # Add a separator row between courses

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Create a temporary file to save the workbook
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="candidate_{candidate.pk}_data.xlsx"'
        return response
